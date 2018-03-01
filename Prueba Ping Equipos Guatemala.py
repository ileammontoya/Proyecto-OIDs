# coding=utf-8
import os
import stat
import openpyxl
import subprocess

wb=openpyxl.load_workbook("QOS_Guatemala.xlsx")
sheet=wb.get_sheet_by_name("Sheet1")

with open(os.devnull, "wb") as limbo:
	for row in range (115,128):
		ip=str(sheet["E"+str(row)].value)
		print ip, row
		result=subprocess.Popen(["ping","-i","0.5","-c", "3", "-n", "-W", "2", ip],
                        stdout=limbo, stderr=limbo).wait()
                if result:
                	sheet.cell(row=row,column=6).value="Unreachable"
                else:
                	sheet.cell(row=row,column=6).value="Active"
wb.save('QOS_Guatemala.xlsx')