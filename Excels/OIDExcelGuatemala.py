# coding=utf-8
import os
import glob
import os
import stat
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment, Side

testcounter=0

titlecolor = PatternFill(fill_type='solid',
						start_color='FF0000',
						end_color='FF0000')
titlefont = Font(size=15,
				color='FFFFFF')

cellalignment = Alignment(horizontal='center',
							vertical='center')

secondcolor = PatternFill(fill_type='solid',
							start_color='000000',
							end_color='000000')

secondfont = Font(color='FFFFFF')

cellcolor = PatternFill(fill_type='lightGrid',
					 	start_color='FFFFFF',
					 	end_color='FFFFFF')

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


#Ciclos redundantes para buscar los indices de cada OID dependiendo de la interfaz donde se encuentra aplicada
#Los primeros dos ciclos y el if dan la relacion de la interfaz y el indice de calidad de servicio
def writeexcel(diccionarios,fila,host):
	global testcounter
	interestingclasses=['Enrutamiento-Senalizacion','EnrutamientoSenalizacion','Datos_VPN_Internet_plus','Ingenieria','Datos_VPN','Video','Aplicaciones-Tiempo-Real', 'class-default']
	qosinterfacepolicy=[]
	qospolicycontrol=0
	#inind==diccionarios[0]
	#inqosind==diccionarios[1]
	#mapas==diccionarios[2]
	#inclaseind==diccionarios[3]
	#counterprev==diccionarios[4]
	#gaugeprev==diccionarios[5]
	#counterpost==diccionarios[6]
	#gaugepost==diccionarios[7]
	#ifipindex==diccionarios[8]
	#ifqosindex==diccionarios[9]
	#queueingindex==diccionarios[10]
	#queueingcurrent==diccionarios[11]
	#queueingmax==diccionarios[12]
	#queueingdiscards==diccionarios[13]

	#Ciclos para juntar Interfaz con la politica que esta aplicada
	#se hicieron en el mismo orden que los ciclos de impresion en excel asi que el orden las politicas e interfaces coindiden en esta lista y en el excel
	for key in diccionarios[0]:
			for item in diccionarios[1]:
				if diccionarios[0][key][0]==diccionarios[1][item][1]:
					for extra in diccionarios[3]:
						if diccionarios[1][item][0] in diccionarios[3][extra][0]:
							for indice in diccionarios[9]:
								if diccionarios[3][extra][1]==diccionarios[9][indice][0]:
									qosinterfacepolicy.append(diccionarios[0][key][1]+' - Policy-Map: '+str(diccionarios[9][indice][1]))
									# test the list in a file
									# my_file=open("oidtest","a")
									# my_file.write(diccionarios[0][key][1]+' - Policy-Map: '+str(diccionarios[9][indice][1]))					
									# my_file.write('\n')
									# my_file.write('\n')
									# my_file.close()



	for key in diccionarios[0]:
		for item in diccionarios[1]:
			#OLD FILTER FOR CONTROL PLANE AND VLANS
			# if diccionarios[0][key][0]==diccionarios[1][item][1] and (diccionarios[0][key][1]=='Control Plane' or 'Vlan' in diccionarios[0][key][1]):
			# 	sheet.cell(row=fila,column=1).value=host+' - '+diccionarios[8]['value1'][0]
			# 	sheet.merge_cells(start_row=fila,start_column=1, end_row=fila, end_column=4)
			# 	sheet.cell(row=fila,column=1).font=titlefont
			# 	sheet.cell(row=fila,column=1).fill=titlecolor
			# 	sheet.cell(row=fila,column=1).alignment=cellalignment
			# 	sheet.row_dimensions[fila].height=20
			# 	sheet.cell(row=fila,column=1).border=thin_border
			# 	sheet.cell(row=fila,column=2).border=thin_border
			# 	sheet.cell(row=fila,column=3).border=thin_border
			# 	sheet.cell(row=fila,column=4).border=thin_border

			# 	fila+=1

			# 	sheet.cell(row=fila,column=1).value=diccionarios[0][key][1]
			# 	qospolicycontrol+=1
			# 	sheet.merge_cells(start_row=fila,start_column=1, end_row=fila, end_column=4)
			# 	sheet.cell(row=fila,column=1).font=titlefont
			# 	sheet.cell(row=fila,column=1).fill=titlecolor
			# 	sheet.cell(row=fila,column=1).alignment=cellalignment
			# 	sheet.row_dimensions[fila].height=20
			# 	sheet.cell(row=fila,column=1).border=thin_border
			# 	sheet.cell(row=fila,column=2).border=thin_border
			# 	sheet.cell(row=fila,column=3).border=thin_border
			# 	sheet.cell(row=fila,column=4).border=thin_border

			# 	fila+=1

			# 	sheet.cell(row=fila,column=1).value='INTERFAZ CONTROL PLANE O VLAN - NO SE INCLUYEN OIDS'
			# 	sheet.merge_cells(start_row=fila,start_column=1, end_row=fila, end_column=4)
			# 	sheet.cell(row=fila,column=1).font=secondfont
			# 	sheet.cell(row=fila,column=1).fill=secondcolor
			# 	sheet.cell(row=fila,column=1).alignment=cellalignment
			# 	sheet.cell(row=fila,column=1).border=thin_border
			# 	sheet.cell(row=fila,column=2).border=thin_border
			# 	sheet.cell(row=fila,column=3).border=thin_border
			# 	sheet.cell(row=fila,column=4).border=thin_border

			# 	fila+=1

			#OLD STATEMENT
			# if diccionarios[0][key][0]==diccionarios[1][item][1] and (diccionarios[0][key][1]!='Control Plane' and 'Vlan' not in diccionarios[0][key][1]):
			if diccionarios[0][key][0]==diccionarios[1][item][1]:

				sheet.cell(row=fila,column=1).value=host+' - '+diccionarios[8]['value1'][0]
				sheet.merge_cells(start_row=fila,start_column=1, end_row=fila, end_column=3)
				sheet.cell(row=fila,column=1).font=titlefont
				sheet.cell(row=fila,column=1).fill=titlecolor
				sheet.cell(row=fila,column=1).alignment=cellalignment
				sheet.row_dimensions[fila].height=20
				sheet.cell(row=fila,column=1).border=thin_border
				sheet.cell(row=fila,column=2).border=thin_border
				sheet.cell(row=fila,column=3).border=thin_border
				# sheet.cell(row=fila,column=4).border=thin_border

				fila+=1

				sheet.cell(row=fila,column=1).value=diccionarios[0][key][1]
				qospolicycontrol+=1
				sheet.merge_cells(start_row=fila,start_column=1, end_row=fila, end_column=3)
				sheet.cell(row=fila,column=1).font=titlefont
				sheet.cell(row=fila,column=1).fill=titlecolor
				sheet.cell(row=fila,column=1).alignment=cellalignment
				sheet.row_dimensions[fila].height=20
				sheet.cell(row=fila,column=1).border=thin_border
				sheet.cell(row=fila,column=2).border=thin_border
				sheet.cell(row=fila,column=3).border=thin_border
				# sheet.cell(row=fila,column=4).border=thin_border

				fila+=1

				sheet.cell(row=fila,column=1).value='Class-Map'
				sheet.cell(row=fila,column=1).font=secondfont
				sheet.cell(row=fila,column=1).fill=secondcolor
				sheet.cell(row=fila,column=1).alignment=cellalignment
				sheet.cell(row=fila,column=1).border=thin_border

				sheet.cell(row=fila,column=3).value='OID'
				sheet.cell(row=fila,column=3).font=secondfont
				sheet.cell(row=fila,column=3).fill=secondcolor
				sheet.cell(row=fila,column=3).alignment=cellalignment
				sheet.cell(row=fila,column=3).border=thin_border

				sheet.cell(row=fila,column=2).value='Descripcion de OID'
				sheet.cell(row=fila,column=2).font=secondfont
				sheet.cell(row=fila,column=2).fill=secondcolor
				sheet.cell(row=fila,column=2).alignment=cellalignment
				sheet.cell(row=fila,column=2).border=thin_border

				# sheet.cell(row=fila,column=4).value='Valor Muestra de cada OID'
				# sheet.cell(row=fila,column=4).font=secondfont
				# sheet.cell(row=fila,column=4).fill=secondcolor
				# sheet.cell(row=fila,column=4).alignment=cellalignment
				# sheet.cell(row=fila,column=4).border=thin_border

				fila+=1

				#los segundos dos ciclos y el if escriben buscan en base a la informacion de QOS e indices de interfaz
				#Imprimen lo mas relevante en el archivo de texto


				corevar=0
				for text in diccionarios[2]:
					for more in diccionarios[3]:
						if diccionarios[1][item][0] == diccionarios[3][more][0].split('.')[0] and diccionarios[2][text][0]==diccionarios[3][more][1]:
							for finalvalue in diccionarios[4]:
								if diccionarios[4][finalvalue][0]==diccionarios[3][more][0]:
									if diccionarios[2][text][1] in interestingclasses: 
										corevar+=1
									sheet.cell(row=fila,column=1).value=diccionarios[2][text][1]
									sheet.cell(row=fila,column=1).fill=cellcolor
									sheet.cell(row=fila,column=1).border=thin_border
									sheet.cell(row=fila,column=3).value='1.3.6.1.4.1.9.9.166.1.15.1.1.6.'+diccionarios[3][more][0]
									sheet.cell(row=fila,column=3).fill=cellcolor
									sheet.cell(row=fila,column=3).border=thin_border
									sheet.cell(row=fila,column=2).value='Contador de Bytes PRE-Politica'
									sheet.cell(row=fila,column=2).fill=cellcolor
									sheet.cell(row=fila,column=2).border=thin_border
									# sheet.cell(row=fila,column=4).value=diccionarios[4][finalvalue][1]
									# sheet.cell(row=fila,column=4).fill=cellcolor
									# sheet.cell(row=fila,column=4).border=thin_border

									fila+=1

									if diccionarios[2][text][1] in interestingclasses:
										corevar+=1
									sheet.cell(row=fila,column=1).value=diccionarios[2][text][1]
									sheet.cell(row=fila,column=1).fill=cellcolor
									sheet.cell(row=fila,column=1).border=thin_border
									sheet.cell(row=fila,column=3).value='1.3.6.1.4.1.9.9.166.1.15.1.1.10.'+diccionarios[3][more][0]
									sheet.cell(row=fila,column=3).fill=cellcolor
									sheet.cell(row=fila,column=3).border=thin_border
									sheet.cell(row=fila,column=2).value='Contador de Bytes POST-Politica'
									sheet.cell(row=fila,column=2).fill=cellcolor
									sheet.cell(row=fila,column=2).border=thin_border
									# sheet.cell(row=fila,column=4).value=diccionarios[6][finalvalue][1]
									# sheet.cell(row=fila,column=4).fill=cellcolor
									# sheet.cell(row=fila,column=4).border=thin_border
									fila+=1

				#SEQUENCE FOR QUEUEING OIDs
				# for text in diccionarios[10]:
				# 	for finalvalue in diccionarios[11]:
				# 		if diccionarios[11][finalvalue][0]==diccionarios[10][text][0] and diccionarios[10][text][0].split('.')[0]==diccionarios[1][item][0]:
				# 			# print diccionarios[10][text][3], diccionarios[12][finalvalue][0], diccionarios[12][finalvalue][1]
				# 			if diccionarios[10][text][3] in interestingclasses: 
				# 					corevar+=1
				# 			sheet.cell(row=fila,column=2).value=diccionarios[10][text][3]
				# 			sheet.cell(row=fila,column=2).fill=cellcolor
				# 			sheet.cell(row=fila,column=2).border=thin_border
				# 			sheet.cell(row=fila,column=3).value='1.3.6.1.4.1.9.9.166.1.18.1.1.2.'+diccionarios[12][finalvalue][0]
				# 			sheet.cell(row=fila,column=3).fill=cellcolor
				# 			sheet.cell(row=fila,column=3).border=thin_border
				# 			sheet.cell(row=fila,column=1).value='Maxima cantidad de paquetes en cola'
				# 			sheet.cell(row=fila,column=1).fill=cellcolor
				# 			sheet.cell(row=fila,column=1).border=thin_border
				# 			sheet.cell(row=fila,column=4).value=diccionarios[12][finalvalue][1]
				# 			sheet.cell(row=fila,column=4).fill=cellcolor
				# 			sheet.cell(row=fila,column=4).border=thin_border

				# 			fila+=1

				# 			# print diccionarios[10][text][3], diccionarios[11][finalvalue][0], diccionarios[11][finalvalue][1]
				# 			if diccionarios[10][text][3] in interestingclasses: 
				# 					corevar+=1
				# 			sheet.cell(row=fila,column=2).value=diccionarios[10][text][3]
				# 			sheet.cell(row=fila,column=2).fill=cellcolor
				# 			sheet.cell(row=fila,column=2).border=thin_border
				# 			sheet.cell(row=fila,column=3).value='1.3.6.1.4.1.9.9.166.1.18.1.1.1.'+diccionarios[11][finalvalue][0]
				# 			sheet.cell(row=fila,column=3).fill=cellcolor
				# 			sheet.cell(row=fila,column=3).border=thin_border
				# 			sheet.cell(row=fila,column=1).value='Actual cantidad de paquetes en cola'
				# 			sheet.cell(row=fila,column=1).fill=cellcolor
				# 			sheet.cell(row=fila,column=1).border=thin_border
				# 			sheet.cell(row=fila,column=4).value=diccionarios[11][finalvalue][1]
				# 			sheet.cell(row=fila,column=4).fill=cellcolor
				# 			sheet.cell(row=fila,column=4).border=thin_border

				# 			fila+=1
				# 			# print diccionarios[10][text][3], diccionarios[13][finalvalue][0], diccionarios[13][finalvalue][1]
				# 			if diccionarios[10][text][3] in interestingclasses: 
				# 					corevar+=1
				# 			sheet.cell(row=fila,column=2).value=diccionarios[10][text][3]
				# 			sheet.cell(row=fila,column=2).fill=cellcolor
				# 			sheet.cell(row=fila,column=2).border=thin_border
				# 			sheet.cell(row=fila,column=3).value='1.3.6.1.4.1.9.9.166.1.18.1.1.5.'+diccionarios[13][finalvalue][0]
				# 			sheet.cell(row=fila,column=3).fill=cellcolor
				# 			sheet.cell(row=fila,column=3).border=thin_border
				# 			sheet.cell(row=fila,column=1).value='Cantidad de Bytes descartados por cola'
				# 			sheet.cell(row=fila,column=1).fill=cellcolor
				# 			sheet.cell(row=fila,column=1).border=thin_border
				# 			sheet.cell(row=fila,column=4).value=diccionarios[13][finalvalue][1]
				# 			sheet.cell(row=fila,column=4).fill=cellcolor
				# 			sheet.cell(row=fila,column=4).border=thin_border

				# 			fila+=1
							# pass
				
				#WHEN INCLUDING QUEUEING OIDs corevar is checked for 32 and 34 rows are subtracted from fila to write Core-MPLS
				if corevar==14:
					sheet.cell(row=fila-16,column=1).value=diccionarios[0][key][1]+' - Core-MPLS'
				for i in range (0,5):
					sheet.cell(row=fila+i,column=1).value=''			
	return fila

#Funcion para escribir los OIDs de cada Class-map en los archivos de texto
def appendoids(archivo,filaultimate):
	#Guarda la informacion de los OID obtenidos en la variable results
	results=[]
	with open(str(archivo), 'r') as inputfile:
		for line in inputfile:
			results.append(line.strip())

	#Genera un "indice" de donde se encuentra la informacion para cada tipo de OID
	#Resultado es una lista donde se expresa cuantas lineas hay en cada OID
	indexlist=[]
	last=0
	for i in range (0, len(results)):
		if "#" in results[i]:
			indexlist.append(i-last-3)
			last=i
	#vacia los diccionarios de cada OID
	hostname={}
	ifindex={}
	ifipindex={}
	ifipindextemp={}
	ifqosindex={}
	ifdir={}
	policymaps={}
	classmaps={}
	ifclassindex={}
	counterprev={}
	gaugeprev={}
	counterpost={}
	gaugepost={}
	parentclassmaps={}
	objecttype={}
	queueingindex={}
	queueingcurrent={}
	queueingmax={}
	queueingdiscards={}


	#Guarda en cada diccionario la informacion mas relevante de cada OID
	#Las keys de cada diccionario empiezan por value1 y aumentan dependiendo de la cantidad de lineas en cada OID
	#El ciclo depende de la dimension de cada archivo de texto
	#Cada "if" se activa con la descripcion que identifica los OID
	for i in range (0,len(results)):
		if "#Hostname" in results[i]:
			temp=[]
			#ciclo usa la cantidad de lineas de acuerdo a su OID
			for x in range (0,indexlist[1]):
				#primer paso temp separa el string en dos usando el simbolo =
				temp=results[i+x+1].split(":")
				if len(temp) >=3:
					#segundo paso guarda el ultimo valor del OID y el valor de la info correspondiente - se limpian los espacios y simbolos extras
					#en caso de problemas de output imprimir temp antes del if para ver su formato
					temp=[temp[0].split(".")[-1].strip(),temp[2].strip().strip('"')]
					hostname["value{0}".format(x+1)]=temp
		if "#Indice de las interfaces" in results[i]:
			temp=[]
			for x in range (0,indexlist[2]):
				temp=results[i+x+1].split(":")
				if len(temp) >=3:
					temp=[temp[0].split(".")[-1].strip(),temp[2].strip().strip('"')]
					ifindex["value{0}".format(x+1)]=temp
		if "#Indice de IPs e interfaces" in results[i]:
			temp=[]
			for x in range (0,indexlist[3]):
				temp=results[i+x+1].split(":")
				if len(temp) >=3:
					temp=[".".join(temp[0].split(".")[-4:]),temp[2].strip().strip('"')]
					ifipindextemp["value{0}".format(x+1)]=temp
		if "#Indice de Calidad de Servicio aplicado a cada interfaz" in results[i]:
			temp=[]
			for x in range (0,indexlist[4]):
				temp=results[i+x+1].split(":")
				if len(temp) >=3:
					temp=[temp[0].split(".")[-1].strip(),temp[2].strip().strip('"')]
					ifqosindex["value{0}".format(x+1)]=temp
		if "#Direccion en la cual se esta aplicando la politica" in results[i]:
			temp=[]
			for x in range (0,indexlist[5]):
				temp=results[i+x+1].split(":")
				if len(temp) >=3:
					temp=[temp[0].split(".")[-1].strip(),temp[2].strip().strip('"')]
					ifdir["value{0}".format(x+1)]=temp
		if "#Policy maps configurados en el equipo" in results[i]:
			temp=[]
			for x in range (0,indexlist[6]):
				temp=results[i+x+1].split(":")
				if len(temp) >=3:
					temp=[temp[0].split(".")[-1].strip(),temp[2].strip().strip('"')]
					policymaps["value{0}".format(x+1)]=temp
		if "#Class maps configurados en el equipo" in results[i]:
			temp=[]
			for x in range (0,indexlist[7]):
				temp=results[i+x+1].split(":")
				if len(temp) >=3:
					temp=[temp[0].split(".")[-1].strip(),temp[2].strip().strip('"')]
					classmaps["value{0}".format(x+1)]=temp
		if "#Indice Parent Classes" in results[i]:
			temp=[]
			for x in range (0,indexlist[8]):
				temp=results[i+x+1].split(":")
				if len(temp) >=3:
					temp=[temp[0].split(".")[-2].strip()+'.'+temp[0].split(".")[-1].strip(),temp[2].strip().strip('"')]
					parentclassmaps["value{0}".format(x+1)]=temp
		if "#Indice usando Class-maps e Interfaces" in results[i]:
			temp=[]
			for x in range (0,indexlist[9]):
				temp=results[i+x+1].split(":")
				if len(temp) >=3:
					temp=[temp[0].split(".")[-2].strip()+'.'+temp[0].split(".")[-1].strip(),temp[2].strip().strip('"')]
					ifclassindex["value{0}".format(x+1)]=temp
		if "#Valores del Contador 64 bits - Previo a ejecutar Politicas" in results[i]:
			temp=[]
			for x in range (0,indexlist[10]):
				temp=results[i+x+1].split(":")
				if len(temp) >=3:
					temp=[temp[0].split(".")[-2].strip()+'.'+temp[0].split(".")[-1].strip(),temp[2].strip().strip('"')]
					counterprev["value{0}".format(x+1)]=temp
		if "#Valores del Gauge32 - Previo a ejecutar Politicas" in results[i]:
			temp=[]
			for x in range (0,indexlist[11]):
				temp=results[i+x+1].split(":")
				if len(temp) >=3:
					temp=[temp[0].split(".")[-2].strip()+'.'+temp[0].split(".")[-1].strip(),temp[2].strip().strip('"')]
					gaugeprev["value{0}".format(x+1)]=temp
		if "#Valores del Contador 64 bits - Despues a ejecutar Politicas" in results[i]:
			temp=[]
			for x in range (0,indexlist[12]):
				temp=results[i+x+1].split(":")
				if len(temp) >=3:
					temp=[temp[0].split(".")[-2].strip()+'.'+temp[0].split(".")[-1].strip(),temp[2].strip().strip('"')]
					counterpost["value{0}".format(x+1)]=temp
		if "#Valores del Gauge32 - Despues a ejecutar Politicas" in results[i]:
			temp=[]
			for x in range (0,indexlist[13]):
				temp=results[i+x+1].split(":")
				if len(temp) >=3:
					temp=[temp[0].split(".")[-2].strip()+'.'+temp[0].split(".")[-1].strip(),temp[2].strip().strip('"')]
					gaugepost["value{0}".format(x+1)]=temp
		if "#Object Type" in results[i]:
			temp=[]
			for x in range (0,indexlist[14]):
				temp=results[i+x+1].split(":")
				if len(temp) >=3:
					temp=[temp[0].split(".")[-2].strip()+'.'+temp[0].split(".")[-1].strip(),temp[2].strip().strip('"')]
					objecttype["value{0}".format(x+1)]=temp
		if "#Queueing current depth" in results[i]:
			temp=[]
			for x in range (0,indexlist[15]):
				temp=results[i+x+1].split(":")
				if len(temp) >=3:
					temp=[temp[0].split(".")[-2].strip()+'.'+temp[0].split(".")[-1].strip(),temp[2].strip().strip('"')]
					queueingcurrent["value{0}".format(x+1)]=temp
		if "#Queueing max depth" in results[i]:
			temp=[]
			for x in range (0,indexlist[16]):
				temp=results[i+x+1].split(":")
				if len(temp) >=3:
					temp=[temp[0].split(".")[-2].strip()+'.'+temp[0].split(".")[-1].strip(),temp[2].strip().strip('"')]
					queueingmax["value{0}".format(x+1)]=temp
		if "#Queueing discards" in results[i]:
			temp=[]
			for x in range (0,indexlist[17]):
				temp=results[i+x+1].split(":")
				if len(temp) >=3:
					temp=[temp[0].split(".")[-2].strip()+'.'+temp[0].split(".")[-1].strip(),temp[2].strip().strip('"')]
					queueingdiscards["value{0}".format(x+1)]=temp


	#Getting only the loopback
	for key in ifindex:
		if ifindex[key][1]=='Loopback0':
			for ip in ifipindextemp:
				if ifindex[key][0]==ifipindextemp[ip][1]:
					ifipindex={'value1':[ifipindextemp[ip][0],ifindex[key][0]]}


	#Ciclos para obtener un indice que incluya informacion relevante de las colas 
	#Como viene presentado en los OID originales no se tiene suficiente infomacion para relacionar los class-maps con cada cola
	r=1
	for key in objecttype:
		for testvar in parentclassmaps:
			if 	objecttype[key][0]==parentclassmaps[testvar][0] and objecttype[key][1]=='queueing':
				for index in ifclassindex:
					if objecttype[key][0]==ifclassindex[index][0]:
						queueingindex["value{0}".format(r)]=[objecttype[key][0], ifclassindex[index][1], objecttype[key][0].split('.')[0]+'.'+parentclassmaps[testvar][1]]
						r+=1

	for key in ifclassindex:
		for testvar in classmaps:
			if classmaps[testvar][0]==ifclassindex[key][1]:
				for more in queueingindex:
					if queueingindex[more][2]==ifclassindex[key][0]:
						queueingindex[more].append(classmaps[testvar][1])


	dicts=[ifindex,ifqosindex,classmaps,ifclassindex,counterprev,gaugeprev,counterpost,gaugepost,ifipindex,policymaps,queueingindex,queueingcurrent,queueingmax,queueingdiscards]


	filaultimate = writeexcel(dicts,filaultimate,hostname['value1'][1])+4
	return filaultimate


#lista los archivos de texto en el directorio y pasa cada uno por la funcion para aplicar 
print 'Starting...'
print
wb=openpyxl.load_workbook("OID_Guatemala.xlsx")
sheet=wb.get_sheet_by_name("Sheet1")
row=1
policyclass=[]
classpolicy=[]
for filename in sorted(glob.glob('*.txt')):
	print filename
	row=appendoids(filename,row)
print
print 'Writing to Excel File'
wb.save('OID_Guatemala.xlsx')

