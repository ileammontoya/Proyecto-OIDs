# coding=utf-8
import os
import glob
import os
import stat
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment, Side
from copy import copy, deepcopy

titlecolor = PatternFill(fill_type='solid',
						start_color='FF0000',
						end_color='FF0000')
titlefont = Font(size=15,
				color='FFFFFF')

cellalignment = Alignment(horizontal='center',
							vertical='center')

secondcolor = PatternFill(fill_type='solid',
							start_color='000000',
							end_color='000000')

secondfont = Font(color='FFFFFF')

cellcolor = PatternFill(fill_type='lightGrid',
					 	start_color='FFFFFF',
					 	end_color='FFFFFF')

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


def CopyCoreBlock(oldrow,newrow):
	for i in range (0,17):

		if i <= 1:
			newexcel.merge_cells(start_row=newrow+i,start_column=1, end_row=newrow+i, end_column=3)
			newexcel.row_dimensions[newrow+i].height=20
		if '10.192' in oldexcel.cell(row=oldrow+i,column=1).value:
			print oldexcel.cell(row=oldrow+i,column=1).value
		newexcel.cell(row=newrow+i,column=1).value=oldexcel.cell(row=oldrow+i,column=1).value
		newexcel.cell(row=newrow+i,column=1).font=oldexcel.cell(row=oldrow+i,column=1).font.copy()
		newexcel.cell(row=newrow+i,column=1).fill=oldexcel.cell(row=oldrow+i,column=1).fill.copy()
		newexcel.cell(row=newrow+i,column=1).border=oldexcel.cell(row=oldrow+i,column=1).border.copy()
		newexcel.cell(row=newrow+i,column=1).alignment=oldexcel.cell(row=oldrow+i,column=1).alignment.copy()

		newexcel.cell(row=newrow+i,column=2).value=oldexcel.cell(row=oldrow+i,column=2).value
		newexcel.cell(row=newrow+i,column=2).font=oldexcel.cell(row=oldrow+i,column=1).font.copy()
		newexcel.cell(row=newrow+i,column=2).fill=oldexcel.cell(row=oldrow+i,column=1).fill.copy()
		newexcel.cell(row=newrow+i,column=2).border=oldexcel.cell(row=oldrow+i,column=1).border.copy()
		newexcel.cell(row=newrow+i,column=2).alignment=oldexcel.cell(row=oldrow+i,column=1).alignment.copy()

		newexcel.cell(row=newrow+i,column=3).value=oldexcel.cell(row=oldrow+i,column=3).value
		newexcel.cell(row=newrow+i,column=3).font=oldexcel.cell(row=oldrow+i,column=1).font.copy()
		newexcel.cell(row=newrow+i,column=3).fill=oldexcel.cell(row=oldrow+i,column=1).fill.copy()
		newexcel.cell(row=newrow+i,column=3).border=oldexcel.cell(row=oldrow+i,column=1).border.copy()
		newexcel.cell(row=newrow+i,column=3).alignment=oldexcel.cell(row=oldrow+i,column=1).alignment.copy()

		# newexcel.cell(row=newrow+i,column=4).value=oldexcel.cell(row=oldrow+i,column=4).value
		# newexcel.cell(row=newrow+i,column=4).font=oldexcel.cell(row=oldrow+i,column=1).font.copy()
		# newexcel.cell(row=newrow+i,column=4).fill=oldexcel.cell(row=oldrow+i,column=1).fill.copy()
		# newexcel.cell(row=newrow+i,column=4).border=oldexcel.cell(row=oldrow+i,column=1).border.copy()
		# newexcel.cell(row=newrow+i,column=4).alignment=oldexcel.cell(row=oldrow+i,column=1).alignment.copy()
	return newrow+i


print 'Starting...'
print
wbold=openpyxl.load_workbook("OID_Guatemala.xlsx")
oldexcel=wbold.get_sheet_by_name("Sheet1")
wbnew=openpyxl.load_workbook("OID_Guatemala_Core-MPLS.xlsx")
newexcel=wbnew.get_sheet_by_name("Sheet1")

invalid=0
orirow=1
formatrow=1
counter=1
oldtitle=''
while orirow < oldexcel.max_row:
	if oldexcel.cell(row=orirow,column=1).value is None:
		invalid+=1
	elif 'Core-MPLS' in oldexcel.cell(row=orirow,column=1).value:
		if oldtitle!=oldexcel.cell(row=orirow-1,column=1).value:
			formatrow+=5
		formatrow=CopyCoreBlock(orirow-1,formatrow)
		print 'Original Row =%s, Equipment Counter=%s, New Row=%s' % (orirow,counter,formatrow)
		oldtitle=oldexcel.cell(row=orirow-1,column=1).value


		formatrow+=1
		counter+=1
	orirow+=1







print oldtitle
print 'Writing to Excel File'
wbnew.save('OID_Guatemala_Core-MPLS.xlsx')

