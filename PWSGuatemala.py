# coding=utf-8
import os
import stat
import openpyxl
wb=openpyxl.load_workbook("QOS_Guatemala.xlsx")
sheet=wb.get_sheet_by_name("Sheet1")
try:
	os.remove("Powershell SNMP Script Guatemala.txt")
except OSError:
	pass
row,column=2,1
salir = sheet.cell(row=row,column=column).value
while salir != None:
	host=sheet["D"+str(row)].value
	ip=sheet["E"+str(row)].value
	community=sheet["G"+str(row)].value
	if str(sheet["F"+str(row)].value)=="Active":
		my_file=open("Powershell SNMP Script Guatemala.txt","a")
		my_file.write('echo "#Hostname" | Out-file C:\users\imontoya\Documents\Guatemala\Host_'+str(host)+'.txt -Append'+"\n")
		my_file.write("snmpwalk -v2c -c "+str(community)+" "+str(ip)+" 1.3.6.1.2.1.1.5 | Out-file C:\users\imontoya\Documents\Guatemala\Host_"+str(host)+".txt -Append"+"\n")
		my_file.write('echo "" "" "#Indice de las interfaces" | Out-file C:\users\imontoya\Documents\Guatemala\Host_'+str(host)+'.txt -Append'+"\n")
		my_file.write("snmpwalk -v2c -c "+str(community)+" "+str(ip)+" 1.3.6.1.2.1.2.2.1.2 | Out-file C:\users\imontoya\Documents\Guatemala\Host_"+str(host)+".txt -Append"+"\n")
		my_file.write('echo "" "" "#Indice de IPs e interfaces" | Out-file C:\users\imontoya\Documents\Guatemala\Host_'+str(host)+'.txt -Append'+"\n")
		my_file.write("snmpwalk -v2c -c "+str(community)+" "+str(ip)+" 1.3.6.1.2.1.4.20.1.2 | Out-file C:\users\imontoya\Documents\Guatemala\Host_"+str(host)+".txt -Append"+"\n")
		my_file.write('echo "" "" "#Indice de Calidad de Servicio aplicado a cada interfaz" | Out-file C:\users\imontoya\Documents\Guatemala\Host_'+str(host)+'.txt -Append'+"\n")
		my_file.write("snmpwalk -v2c -c "+str(community)+" "+str(ip)+" 1.3.6.1.4.1.9.9.166.1.1.1.1.4 | Out-file C:\users\imontoya\Documents\Guatemala\Host_"+str(host)+".txt -Append"+"\n")
		my_file.write('echo "" "" "#Direccion en la cual se esta aplicando la politica" | Out-file C:\users\imontoya\Documents\Guatemala\Host_'+str(host)+'.txt -Append'+"\n")
		my_file.write("snmpwalk -v2c -c "+str(community)+" "+str(ip)+" 1.3.6.1.4.1.9.9.166.1.1.1.1.3 | Out-file C:\users\imontoya\Documents\Guatemala\Host_"+str(host)+".txt -Append"+"\n")
		my_file.write('echo "" "" "#Policy maps configurados en el equipo" | Out-file C:\users\imontoya\Documents\Guatemala\Host_'+str(host)+'.txt -Append'+"\n")
		my_file.write("snmpwalk -v2c -c "+str(community)+" "+str(ip)+" 1.3.6.1.4.1.9.9.166.1.6.1.1.1 | Out-file C:\users\imontoya\Documents\Guatemala\Host_"+str(host)+".txt -Append"+"\n")
		my_file.write('echo "" "" "#Class maps configurados en el equipo" | Out-file C:\users\imontoya\Documents\Guatemala\Host_'+str(host)+'.txt -Append'+"\n")
		my_file.write("snmpwalk -v2c -c "+str(community)+" "+str(ip)+" 1.3.6.1.4.1.9.9.166.1.7.1.1.1 | Out-file C:\users\imontoya\Documents\Guatemala\Host_"+str(host)+".txt -Append"+"\n")
		my_file.write('echo "" "" "#Indice usando Class-maps e Interfaces" | Out-file C:\users\imontoya\Documents\Guatemala\Host_'+str(host)+'.txt -Append'+"\n")
		my_file.write("snmpwalk -v2c -c "+str(community)+" "+str(ip)+" 1.3.6.1.4.1.9.9.166.1.5.1.1.2 | Out-file C:\users\imontoya\Documents\Guatemala\Host_"+str(host)+".txt -Append"+"\n")
		my_file.write('echo "" "" "#Indice Parent Classes" | Out-file C:\users\imontoya\Documents\Guatemala\Host_'+str(host)+'.txt -Append'+"\n")
		my_file.write("snmpwalk -v2c -c "+str(community)+" "+str(ip)+" 1.3.6.1.4.1.9.9.166.1.5.1.1.4 | Out-file C:\users\imontoya\Documents\Guatemala\Host_"+str(host)+".txt -Append"+"\n")
		my_file.write('echo "" "" "#Valores del Contador 64 bits - Previo a ejecutar Politicas" | Out-file C:\users\imontoya\Documents\Guatemala\Host_'+str(host)+'.txt -Append'+"\n")
		my_file.write("snmpwalk -v2c -c "+str(community)+" "+str(ip)+" 1.3.6.1.4.1.9.9.166.1.15.1.1.6 | Out-file C:\users\imontoya\Documents\Guatemala\Host_"+str(host)+".txt -Append"+"\n")
		my_file.write('echo "" "" "#Valores del Gauge32 - Previo a ejecutar Politicas" | Out-file C:\users\imontoya\Documents\Guatemala\Host_'+str(host)+'.txt -Append'+"\n")
		my_file.write("snmpwalk -v2c -c "+str(community)+" "+str(ip)+" 1.3.6.1.4.1.9.9.166.1.15.1.1.7 | Out-file C:\users\imontoya\Documents\Guatemala\Host_"+str(host)+".txt -Append"+"\n")
		my_file.write('echo "" "" "#Valores del Contador 64 bits - Despues a ejecutar Politicas" | Out-file C:\users\imontoya\Documents\Guatemala\Host_'+str(host)+'.txt -Append'+"\n")
		my_file.write("snmpwalk -v2c -c "+str(community)+" "+str(ip)+" 1.3.6.1.4.1.9.9.166.1.15.1.1.10 | Out-file C:\users\imontoya\Documents\Guatemala\Host_"+str(host)+".txt -Append"+"\n")
		my_file.write('echo "" "" "#Valores del Gauge32 - Despues a ejecutar Politicas" | Out-file C:\users\imontoya\Documents\Guatemala\Host_'+str(host)+'.txt -Append'+"\n")
		my_file.write("snmpwalk -v2c -c "+str(community)+" "+str(ip)+" 1.3.6.1.4.1.9.9.166.1.15.1.1.11 | Out-file C:\users\imontoya\Documents\Guatemala\Host_"+str(host)+".txt -Append"+"\n")
		my_file.write('echo "" "" "#Object Type" | Out-file C:\users\imontoya\Documents\Guatemala\Host_'+str(host)+'.txt -Append'+"\n")
		my_file.write("snmpwalk -v2c -c "+str(community)+" "+str(ip)+" 1.3.6.1.4.1.9.9.166.1.5.1.1.3 | Out-file C:\users\imontoya\Documents\Guatemala\Host_"+str(host)+".txt -Append"+"\n")
		my_file.write('echo "" "" "#Queueing current depth" | Out-file C:\users\imontoya\Documents\Guatemala\Host_'+str(host)+'.txt -Append'+"\n")
		my_file.write("snmpwalk -v2c -c "+str(community)+" "+str(ip)+" 1.3.6.1.4.1.9.9.166.1.18.1.1.1 | Out-file C:\users\imontoya\Documents\Guatemala\Host_"+str(host)+".txt -Append"+"\n")
		my_file.write('echo "" "" "#Queueing max depth" | Out-file C:\users\imontoya\Documents\Guatemala\Host_'+str(host)+'.txt -Append'+"\n")
		my_file.write("snmpwalk -v2c -c "+str(community)+" "+str(ip)+" 1.3.6.1.4.1.9.9.166.1.18.1.1.2 | Out-file C:\users\imontoya\Documents\Guatemala\Host_"+str(host)+".txt -Append"+"\n")
		my_file.write('echo "" "" "#Queueing discards" | Out-file C:\users\imontoya\Documents\Guatemala\Host_'+str(host)+'.txt -Append'+"\n")
		my_file.write("snmpwalk -v2c -c "+str(community)+" "+str(ip)+" 1.3.6.1.4.1.9.9.166.1.18.1.1.5 | Out-file C:\users\imontoya\Documents\Guatemala\Host_"+str(host)+".txt -Append"+"\n")
		my_file.write('echo "" "" "#End" | Out-file C:\users\imontoya\Documents\Guatemala\Host_'+str(host)+'.txt -Append'+"\n")
		my_file.write('\n')
		my_file.write('\n')
		my_file.close()
	else:
		inactivos=open("Equipos Guatemala Inactivos.txt","a")
		inactivos.write(str(host)+" - "+str(ip)+" \n")
		inactivos.close()
	row+=1
	salir = sheet.cell(row=row,column=column).value
	print(row,salir)
